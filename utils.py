"""
Módulo de utilitários da automação RPA - Google Maps.

Contém funções auxiliares para:
  - Configuração do sistema de logging (arquivo + console)
  - Persistência dos dados coletados em formato JSON
  - Geração de planilha Excel formatada (.xlsx)
"""

import json
import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config


# ==========================================================================
# Logging
# ==========================================================================

def setup_logger(name: str = "rpa_google_maps") -> logging.Logger:
    """
    Configura e retorna o logger principal da aplicação.

    Cria dois handlers:
      - FileHandler: grava em automation.log com nível DEBUG (tudo).
      - StreamHandler: exibe no console com nível INFO (eventos normais).

    Cada entrada de log contém data/hora, nível e mensagem.

    Args:
        name: Nome do logger.

    Returns:
        logging.Logger: Logger configurado e pronto para uso.
    """
    logger = logging.getLogger(name)
    logger.setLevel(logging.DEBUG)

    # Formato padrão: "2025-01-28 14:30:00 [INFO] Mensagem aqui"
    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    # Handler para arquivo — registra tudo (DEBUG e acima)
    file_handler = logging.FileHandler(config.LOG_FILE, encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    # Handler para console — registra eventos normais (INFO e acima)
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    # Evita duplicação de handlers caso o logger já tenha sido configurado
    if not logger.handlers:
        logger.addHandler(file_handler)
        logger.addHandler(console_handler)

    return logger


# ==========================================================================
# Persistência JSON
# ==========================================================================

def save_json(data: list, filepath: str = None) -> None:
    """
    Salva a lista de estabelecimentos em um arquivo JSON.

    O arquivo é gerado com encoding UTF-8 e indentação de 4 espaços
    para facilitar a leitura.

    Args:
        data: Lista de dicionários com os dados dos estabelecimentos.
        filepath: Caminho do arquivo. Padrão: config.OUTPUT_JSON.

    Raises:
        IOError: Se houver falha ao gravar o arquivo.
    """
    if filepath is None:
        filepath = config.OUTPUT_JSON

    logger = logging.getLogger("rpa_google_maps")

    try:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        logger.info(f"Arquivo JSON salvo com sucesso: {filepath}")
    except IOError as e:
        logger.error(f"Falha ao salvar arquivo JSON: {e}")
        raise


# ==========================================================================
# Geração de planilha Excel
# ==========================================================================

def generate_excel(data: list, filepath: str = None) -> None:
    """
    Gera uma planilha Excel (.xlsx) formatada a partir dos dados coletados.

    Características da planilha:
      - Cabeçalhos em negrito com fundo azul e texto branco.
      - Bordas finas em todas as células com dados.
      - Largura das colunas ajustada automaticamente ao conteúdo.
      - Primeira linha congelada para manter os cabeçalhos visíveis.

    Args:
        data: Lista de dicionários com os dados dos estabelecimentos.
        filepath: Caminho do arquivo. Padrão: config.OUTPUT_EXCEL.

    Raises:
        Exception: Se houver falha ao gerar ou salvar a planilha.
    """
    if filepath is None:
        filepath = config.OUTPUT_EXCEL

    logger = logging.getLogger("rpa_google_maps")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Estabelecimentos"

        # Definição das colunas
        headers = ["Nome", "Tipo", "Nota", "Avaliações", "Endereço"]

        # --- Estilos ---
        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # --- Escreve os cabeçalhos ---
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # --- Escreve os dados ---
        for row_idx, item in enumerate(data, start=2):
            values = [
                item.get("nome", ""),
                item.get("tipo", ""),
                item.get("nota", ""),
                item.get("avaliacoes", ""),
                item.get("endereco", ""),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = cell_alignment

        # --- Ajusta largura das colunas automaticamente ---
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            # Calcula a largura máxima considerando cabeçalho e dados
            max_length = len(header)
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            # Margem extra para boa legibilidade
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

        # Congela a primeira linha para que os cabeçalhos fiquem sempre visíveis
        ws.freeze_panes = "A2"

        wb.save(filepath)
        logger.info(f"Planilha Excel salva com sucesso: {filepath}")

    except Exception as e:
        logger.error(f"Falha ao gerar planilha Excel: {e}")
        raise
